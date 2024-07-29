from osgeo import gdal
import numpy as np
import os
from openpyxl import Workbook,load_workbook
import datetime
import pandas as pd

def LoadData(filename):
    file = gdal.Open(filename)
    if file == None:
        print(filename + " can't be opened!")
        return
    nb = file.RasterCount

    L = []
    for i in range(1, nb + 1):
        band = file.GetRasterBand(i)
        background = band.GetNoDataValue()
        data = band.ReadAsArray()
        data = data.astype(np.float32)
        index = np.where(data == background)
        data[index] = 0
        L.append(data)
    data = np.stack(L,0)
    if nb == 1:
        data = data[0,:,:]
    return data


def WriteTiff(im_data, inputdir, path):
    raster = gdal.Open(inputdir)
    im_width = raster.RasterXSize  # 栅格矩阵的列数
    im_height = raster.RasterYSize  # 栅格矩阵的行数
    im_bands = raster.RasterCount  # 波段数
    im_geotrans = raster.GetGeoTransform()  # 获取仿射矩阵信息
    im_proj = raster.GetProjection()  # 获取投影信息

    if 'int8' in im_data.dtype.name:
        datatype = gdal.GDT_Byte
    elif 'int16' in im_data.dtype.name:
        datatype = gdal.GDT_UInt16
    else:
        datatype = gdal.GDT_Float32

    if len(im_data.shape) == 3:
        im_bands, im_height, im_width = im_data.shape
    elif len(im_data.shape) == 2:
        im_data = np.array([im_data])
    else:
        im_bands, (im_height, im_width) = 1, im_data.shape
        # 创建文件
    driver = gdal.GetDriverByName("GTiff")
    dataset = driver.Create(path, im_width, im_height, im_bands, datatype)
    if (dataset != None):
        dataset.SetGeoTransform(im_geotrans)  # 写入仿射变换参数
        dataset.SetProjection(im_proj)  # 写入投影
    for i in range(im_bands):
        dataset.GetRasterBand(i + 1).WriteArray(im_data[i])
    del dataset

def normalization(data):
    _range = np.max(data) - np.min(data)
    return (data - np.min(data)) / _range

# 精度评价部分
def getOAKappe(pred_arr, true_arr):
    OA_kappa_arr = pred_arr * 10 + true_arr

    TP = np.size(np.where(OA_kappa_arr == 11))  # 结果、实际都是城市
    FP = np.size(np.where(OA_kappa_arr == 10))  # 结果是城市，实际不是城市
    FN = np.size(np.where(OA_kappa_arr == 1))  # 结果不是，实际是城市
    TN = np.size(np.where(OA_kappa_arr == 0))  # 结果，实际都不是城市
    TotalNum = TP + FP + FN + TN
    OverAccury = (TP + TN) / TotalNum
    kappa = (float((TP + TN) * TotalNum) - float(((FN + TN) * (FP + TN) + (FN + TP) * (FP + TP)))) / (
            float(TotalNum * TotalNum) - float(((FN + TN) * (FP + TN) + (FN + TP) * (FP + TP))))
    return OverAccury, kappa


def getFOM(true_arr, pred_arr, startYear_arr):
    changeTrue = true_arr - startYear_arr
    changePred = pred_arr - startYear_arr
    A = np.sum(np.where((changeTrue == 1) & (changePred == 0)))
    B = np.sum(np.where((changeTrue == 1) & (changePred == 1)))
    C = np.sum(np.where((changeTrue == 0) & (changePred == 1)))
    return B / (A + B + C)

path1 = "../Unet_result/Suitability_1_2015_lb_1.tif"
path2 = "../LSTMdata/LSTM2015_nor1.tif"
arr1 = LoadData(path1)
arr2 = LoadData(path2)

startyear = 2015
endyear = 2020

df = pd.read_excel("../cityCount/count{}_{}.xlsx".format(str(startyear), str(endyear)))
arr_city = LoadData("../inputdata/boundary/cityBoundary.tif")

outputpath = "../urbanExpansion/cities"
save_file = outputpath + "/参数和精度cities.xlsx"
if os.path.isfile(save_file):
    os.remove(save_file)
workbook = Workbook()
workbook.save(filename=save_file)

for weight in range(0,11):
    for city_id in np.unique(arr_city):
        if city_id != 0:
            im_data = weight * 0.1 * arr1 + (10 - weight) * 0.1 * arr2
            im_data = normalization(im_data)

            arr_boundary = LoadData("../inputdata/boundary/boundary.tif")
            startpath = "../origindata/urban/urban3/urban{}.tif".format(str(startyear))
            start_arr = LoadData(startpath)
            start_arr[np.where(arr_boundary == 0)] = np.nan
            end_arr = LoadData("../origindata/urban/urban3/urban{}.tif".format(str(endyear)))
            end_arr[np.where(arr_boundary == 0)] = np.nan
            endYearCount = np.sum(end_arr == 1)
            PiexlNumber = endYearCount - np.sum(start_arr == 1)

            developmentProbability_arr = im_data
            developmentProbability_arr = developmentProbability_arr * LoadData("../LUSD/conservation/c" + str(endyear) + ".tif")
            developmentProbability_arr[np.where(arr_boundary == 0)] = 0

            developmentProbability_arr[np.where(start_arr == 1)] = 0
            raster = gdal.Open(startpath)
            nRows = raster.RasterYSize  # 行数
            nCols = raster.RasterXSize  # 列数

            # 按照城市分配
            arr_p = np.copy(developmentProbability_arr)
            arr_p[np.where(arr_city != city_id)] = 0
            counts = int(df[df['ID'] == city_id]["Count"])
            Pk = arr_p.reshape(-1)
            start_arr = start_arr.reshape(-1)
            PkSort_R_Index = np.argsort(-Pk)  # 降序索引
            start_arr[PkSort_R_Index[0:counts]] = 1  # 提取概率最高的赋值为城市[0:PkSort_R_Index]] = 1  #提取概率最高的赋值为城市
            #---------------------------

            start_arr = np.array(start_arr).reshape([nRows, nCols])
            start_arr1 = np.copy(start_arr)
            end_arr1 =np.copy(end_arr)
            start_arr1[np.where(arr_city != city_id)] = 3
            end_arr1[np.where(arr_city != city_id)] = 3

            OverAccury, kappa = getOAKappe(start_arr1, end_arr1)
            start_arr_ture = LoadData(startpath)
            start_arr_ture[np.where(arr_boundary == 0)] = np.nan
            Fom = getFOM(end_arr1, start_arr1, start_arr_ture)
            start_arr[np.where(arr_boundary == 0)] = 0


            print("city,OverAccury,kappa,Fom:" + str(city_id) + str(OverAccury) + " " + str(kappa) + " " + str(Fom))
            if not os.path.exists(outputpath):
                os.makedirs(outputpath)

            wb = load_workbook(save_file)
            if str(city_id) not in wb.sheetnames:
                ws = wb.create_sheet(str(city_id))
                ws=wb[str(city_id)]
                ws.append(["Time", "Name","cityID", "OverAccury", "Kappa", "FOM"])
                wb.save(filename=save_file)

            workbook = load_workbook(save_file)
            now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            worksheet1 = workbook[str(city_id)]
            worksheet1.append([now, "plan3_" + str(weight), city_id,OverAccury, kappa, Fom])
            workbook.save(filename=save_file)

    ResultFile = outputpath + "/plan3_w" + str(weight) + ".tif"

    WriteTiff(start_arr, startpath, ResultFile) # 输出模拟城市
    WriteTiff(developmentProbability_arr, startpath, outputpath + "/plan3_w" + str(weight) + "_sp.tif")# 输出加权后概率

df = pd.read_excel("../cityCount/cityid_name.xls")
with pd.ExcelWriter(outputpath + '/各城市精度.xlsx') as writer:
    for cityid in range(1,14):
        data_test = pd.read_excel(save_file,sheet_name=str(float(cityid)))
        df_1 = data_test.sort_values('FOM',ascending=False)

        sheetname =df[df['id'] == cityid]["市"]
        df_1.to_excel(writer,sheet_name = sheetname.values[0],index=False)
writer.save()
