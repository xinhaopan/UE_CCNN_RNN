# -*- coding: utf-8 -*-
"""
Created on Mon Jan 10 20:05:06 2022

@author: Xhpan
"""

from osgeo import gdal
import numpy as np
import os
from openpyxl import Workbook,load_workbook
import datetime
import pandas as pd

def LoadData(filename):
    file = gdal.Open(filename)
    if file == None:
        print(filename + " can't be opened!")
        return
    nb = file.RasterCount

    L = []
    for i in range(1, nb + 1):
        band = file.GetRasterBand(i)
        background = band.GetNoDataValue()
        data = band.ReadAsArray()
        data = data.astype(np.float32)
        index = np.where(data == background)
        data[index] = 0
        L.append(data)
    data = np.stack(L,0)
    if nb == 1:
        data = data[0,:,:]
    
    return data

def WriteTiff(im_data,inputdir, path):
    raster = gdal.Open(inputdir)
    im_width = raster.RasterXSize #栅格矩阵的列数
    im_height = raster.RasterYSize #栅格矩阵的行数
    im_bands = raster.RasterCount #波段数
    im_geotrans = raster.GetGeoTransform()#获取仿射矩阵信息
    im_proj = raster.GetProjection()#获取投影信息

    
    if 'int8' in im_data.dtype.name:
        datatype = gdal.GDT_Byte
    elif 'int16' in im_data.dtype.name:
        datatype = gdal.GDT_UInt16
    else:
        datatype = gdal.GDT_Float32

    if len(im_data.shape) == 3:
        im_bands, im_height, im_width = im_data.shape
    elif len(im_data.shape) == 2:
        im_data = np.array([im_data])
    else:
        im_bands, (im_height, im_width) = 1, im_data.shape
        # 创建文件
    driver = gdal.GetDriverByName("GTiff")
    dataset = driver.Create(path, im_width, im_height, im_bands, datatype)
    if (dataset != None):
        dataset.SetGeoTransform(im_geotrans)  # 写入仿射变换参数
        dataset.SetProjection(im_proj)  # 写入投影
    for i in range(im_bands):
        dataset.GetRasterBand(i + 1).WriteArray(im_data[i])
    del dataset
    
# 精度评价部分
def getOAKappe(pred_arr, true_arr):
    OA_kappa_arr = pred_arr * 10 + true_arr

    TP = np.size(np.where(OA_kappa_arr == 11))  # 结果、实际都是城市
    FP = np.size(np.where(OA_kappa_arr == 10))  # 结果是城市，实际不是城市
    FN = np.size(np.where(OA_kappa_arr == 1))  # 结果不是，实际是城市
    TN = np.size(np.where(OA_kappa_arr == 0))  # 结果，实际都不是城市
    TotalNum = TP + FP + FN + TN
    OverAccury = (TP + TN) / TotalNum
    kappa = (float((TP + TN) * TotalNum) - float(((FN + TN) * (FP + TN) + (FN + TP) * (FP + TP)))) / (
            float(TotalNum * TotalNum) - float(((FN + TN) * (FP + TN) + (FN + TP) * (FP + TP))))
    return OverAccury, kappa


def getFOM(true_arr, pred_arr, startYear_arr):
    changeTrue = true_arr - startYear_arr
    changePred = pred_arr - startYear_arr
    A = np.sum(np.where((changeTrue == 1) & (changePred == 0)))
    B = np.sum(np.where((changeTrue == 1) & (changePred == 1)))
    C = np.sum(np.where((changeTrue == 0) & (changePred == 1)))
    return B / (A + B + C)


startyear = 2015
endyear = 2020

plans = ["plan1","plan2","plan3"]
for plan in plans:
    if plan == "plan1":
        filepath = "../Unet_result/Suitability_1_2015_lb_1.tif"
    elif plan == "plan2":
        filepath = "../LSTMdata/LSTM2015_nor1.tif"
    elif plan == "plan3":
        filepath = "../Unet_result/suitability_plan3.tif"

    if not os.path.exists("../urbanExpansion"):
        os.makedirs("../urbanExpansion")

    ResultFile = "../urbanExpansion/" + plan + ".tif"

    arr_boundary = LoadData("../inputdata/boundary/boundary.tif")
    startpath = "../origindata/urban/urban3/urban{}.tif".format(str(startyear))
    start_arr = LoadData(startpath)
    start_arr[np.where(arr_boundary == 0)] = np.nan
    end_arr = LoadData("../origindata/urban/urban3/urban{}.tif".format(str(endyear)))
    end_arr[np.where(arr_boundary == 0)] = np.nan
    endYearCount = np.sum(end_arr == 1)
    PiexlNumber = endYearCount - np.sum(start_arr == 1)

    developmentProbability_arr = LoadData(filepath)
    developmentProbability_arr = developmentProbability_arr * LoadData("../LUSD/conservation/c" + str(endyear) + ".tif")
    developmentProbability_arr[np.where(arr_boundary == 0)] = 0
    WriteTiff(developmentProbability_arr, startpath, "../urbanExpansion/" + plan + "_sp.tif")

    developmentProbability_arr[np.where(start_arr == 1)] = 0
    raster = gdal.Open(startpath)
    nRows = raster.RasterYSize  # 行数
    nCols = raster.RasterXSize  # 列数

    df = pd.read_excel("../cityCount/count{}_{}.xlsx".format(str(startyear),str(endyear)))
    arr_city = LoadData("../inputdata/boundary/cityBoundary.tif")
    for city_id in np.unique(arr_city):
        if city_id != 0:
            arr_p = np.copy(developmentProbability_arr)
            arr_p[np.where(arr_city != city_id)] = 0
            counts = int(df[df['ID'] == city_id]["Count"])

            Pk = arr_p.reshape(-1)
            start_arr = start_arr.reshape(-1)
            PkSort_R_Index = np.argsort(-Pk)  #降序索引
            start_arr[PkSort_R_Index[0:counts]] = 1  #提取概率最高的赋值为城市[0:PkSort_R_Index]] = 1  #提取概率最高的赋值为城市

    start_arr = np.array(start_arr).reshape([nRows, nCols])
    OverAccury, kappa = getOAKappe(start_arr, end_arr)
    start_arr1 = LoadData(startpath)
    start_arr1[np.where(arr_boundary == 0)] = np.nan
    Fom = getFOM(end_arr, start_arr,start_arr1)
    start_arr[np.where(arr_boundary == 0)] = 0
    WriteTiff(start_arr,startpath, ResultFile)

    print("OverAccury,kappa,Fom:" + str(OverAccury) + " " + str(kappa) + " " + str(Fom))

    save_file = "../urbanExpansion/参数和精度.xlsx"
    if not os.path.isfile(save_file):
        workbook = Workbook()
        worksheet1 = workbook.active
        worksheet1.append(["Time","Name","OverAccury","Kappa","FOM"])
        workbook.save(filename=save_file)

    workbook = load_workbook(save_file)
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    worksheet1 = workbook.active
    worksheet1.append([now,plan,OverAccury,kappa,Fom])
    workbook.save(filename=save_file)


